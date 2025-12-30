# Off-Hours Log Report Email Service
# Auto-generates and sends email reports for logs during 18h-6h

import os
import smtplib
import logging
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import mysql.connector
from jinja2 import Template

logger = logging.getLogger(__name__)

# Email configuration from environment
SMTP_HOST = os.getenv('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_EMAIL = os.getenv('SMTP_EMAIL', '')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')
REPORT_RECIPIENTS = os.getenv('REPORT_RECIPIENTS', '').split(',')

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'mariadb'),
    'user': os.getenv('DB_USER', 'openemr'),
    'password': os.getenv('DB_PASSWORD', 'Openemr!123'),
    'database': os.getenv('DB_NAME', 'ehr_core'),
    'charset': 'utf8mb4'
}

# Timezone offset for Vietnam (UTC+7)
TIMEZONE_OFFSET = 7


def get_off_hours_logs(from_time: datetime, to_time: datetime) -> List[Dict[str, Any]]:
    """
    Get all logs from access_logs table within the specified time range.
    Typically used for off-hours: 18:00 -> 06:00 next day
    """
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
        
        sql = """
            SELECT 
                id, timestamp, user_id, actor_name, role, action, 
                operation, method, status, purpose, uri, ip_address,
                user_agent, log_type, patient_name, patient_code, details
            FROM access_logs
            WHERE timestamp >= %s AND timestamp < %s
            ORDER BY timestamp DESC
            LIMIT 500
        """
        cursor.execute(sql, (from_time, to_time))
        logs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return logs
    except Exception as e:
        logger.error(f"Error fetching off-hours logs: {e}")
        return []


def categorize_logs(logs: List[Dict]) -> Dict[str, Any]:
    """Categorize logs into violations, successes, warnings"""
    violations = []
    successes = []
    warnings = []
    users_active = set()
    ip_addresses = set()
    
    for log in logs:
        action = (log.get('action') or '').lower()
        status = log.get('status') or 0
        log_type = (log.get('log_type') or '').upper()
        user = log.get('user_id') or log.get('actor_name') or 'unknown'
        ip = log.get('ip_address') or ''
        
        users_active.add(user)
        if ip:
            ip_addresses.add(ip)
        
        # Categorize based on status and action
        if 'brute' in action or 'attack' in action or log_type == 'SECURITY_ALERT':
            violations.append(log)
        elif status >= 400 or 'thất bại' in action or 'failed' in action or 'locked' in action:
            warnings.append(log)
        else:
            successes.append(log)
    
    return {
        'violations': violations,
        'successes': successes,
        'warnings': warnings,
        'users_active': list(users_active),
        'ip_addresses': list(ip_addresses),
        'total': len(logs)
    }


HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body { font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px; }
        h2 { color: #1a237e; margin-bottom: 5px; }
        .time-range { color: #666; font-size: 14px; margin-bottom: 20px; }
        table { width: 100%; border-collapse: collapse; margin: 15px 0; }
        th, td { padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #f5f5f5; font-weight: 600; }
        .number { font-weight: bold; font-size: 18px; }
        .danger { color: #c62828; }
        .warning { color: #ef6c00; }
        .success { color: #2e7d32; }
        .footer { margin-top: 20px; padding-top: 15px; border-top: 1px solid #ddd; color: #666; font-size: 12px; }
        .note { background: #e3f2fd; padding: 10px; border-radius: 4px; margin: 15px 0; font-size: 13px; }
    </style>
</head>
<body>
    <h2>🌙 Báo Cáo Log Ngoài Giờ Làm Việc</h2>
    <p class="time-range">Khung giờ: {{ from_time }} → {{ to_time }}</p>
    
    <table>
        <tr>
            <th>Thống kê</th>
            <th>Số lượng</th>
        </tr>
        <tr>
            <td>Tổng logs</td>
            <td class="number">{{ stats.total }}</td>
        </tr>
        <tr>
            <td>🚨 Vi phạm / Tấn công</td>
            <td class="number danger">{{ stats.violations | length }}</td>
        </tr>
        <tr>
            <td>⚠️ Cảnh báo</td>
            <td class="number warning">{{ stats.warnings | length }}</td>
        </tr>
        <tr>
            <td>✅ Thành công</td>
            <td class="number success">{{ stats.successes | length }}</td>
        </tr>
        <tr>
            <td>👤 Users hoạt động</td>
            <td class="number">{{ stats.users_active | length }}</td>
        </tr>
        <tr>
            <td>🌐 IP Addresses</td>
            <td class="number">{{ stats.ip_addresses | length }}</td>
        </tr>
    </table>
    
    <div class="note">
        📎 <strong>Chi tiết đầy đủ</strong> trong file Excel đính kèm.
    </div>
    
    <div class="footer">
        📧 Báo cáo tự động từ SIEM Dashboard | {{ generated_at }}
    </div>
</body>
</html>
"""


def generate_html_report(logs: List[Dict], from_time: datetime, to_time: datetime) -> str:
    """Generate HTML email content from logs"""
    stats = categorize_logs(logs)
    
    template = Template(HTML_TEMPLATE)
    html = template.render(
        stats=stats,
        from_time=from_time.strftime('%d/%m/%Y %H:%M'),
        to_time=to_time.strftime('%d/%m/%Y %H:%M'),
        generated_at=datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    )
    
    return html


def generate_excel_report(logs: List[Dict], from_time: datetime, to_time: datetime) -> bytes:
    """Generate Excel file with all log details"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Off-Hours Logs"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Thời gian', 'User', 'Role', 'Hành động', 'Status', 'IP Address', 'Log Type', 'URI', 'Chi tiết']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Column widths
        ws.column_dimensions['A'].width = 20  # Timestamp
        ws.column_dimensions['B'].width = 15  # User
        ws.column_dimensions['C'].width = 12  # Role
        ws.column_dimensions['D'].width = 50  # Action
        ws.column_dimensions['E'].width = 10  # Status
        ws.column_dimensions['F'].width = 15  # IP
        ws.column_dimensions['G'].width = 20  # Log Type
        ws.column_dimensions['H'].width = 30  # URI
        ws.column_dimensions['I'].width = 40  # Details
        
        # Data rows
        for row_num, log in enumerate(logs, 2):
            timestamp = log.get('timestamp')
            if hasattr(timestamp, 'strftime'):
                timestamp = timestamp.strftime('%d/%m/%Y %H:%M:%S')
            
            ws.cell(row=row_num, column=1, value=str(timestamp or ''))
            ws.cell(row=row_num, column=2, value=log.get('user_id') or log.get('actor_name') or '')
            ws.cell(row=row_num, column=3, value=log.get('role') or '')
            ws.cell(row=row_num, column=4, value=log.get('action') or '')
            ws.cell(row=row_num, column=5, value=str(log.get('status') or ''))
            ws.cell(row=row_num, column=6, value=log.get('ip_address') or '')
            ws.cell(row=row_num, column=7, value=log.get('log_type') or '')
            ws.cell(row=row_num, column=8, value=log.get('uri') or '')
            
            # Details (truncated)
            details = log.get('details') or ''
            if isinstance(details, dict):
                import json
                details = json.dumps(details, ensure_ascii=False)[:200]
            ws.cell(row=row_num, column=9, value=str(details)[:200] if details else '')
            
            # Apply border
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = border
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Tổng kết")
        stats = categorize_logs(logs)
        
        summary_data = [
            ['BÁO CÁO LOG NGOÀI GIỜ LÀM VIỆC', ''],
            ['', ''],
            ['Khung giờ', f"{from_time.strftime('%d/%m/%Y %H:%M')} → {to_time.strftime('%d/%m/%Y %H:%M')}"],
            ['', ''],
            ['THỐNG KÊ', 'SỐ LƯỢNG'],
            ['Tổng logs', stats['total']],
            ['Vi phạm / Tấn công', len(stats['violations'])],
            ['Cảnh báo', len(stats['warnings'])],
            ['Thành công', len(stats['successes'])],
            ['', ''],
            ['Users hoạt động', ', '.join(stats['users_active'][:10])],
            ['IP Addresses', ', '.join(stats['ip_addresses'][:10])],
        ]
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_num, column=1, value=label)
            ws_summary.cell(row=row_num, column=2, value=value)
            if row_num == 1:
                ws_summary.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            if row_num == 5:
                ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
                ws_summary.cell(row=row_num, column=2).font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return None


def send_email_report(subject: str, html_content: str, recipients: List[str], excel_data: bytes = None, excel_filename: str = None) -> bool:
    """Send HTML email via Gmail SMTP with optional Excel attachment"""
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        logger.error("SMTP_EMAIL or SMTP_PASSWORD not configured")
        return False
    
    if not recipients or not recipients[0]:
        logger.error("No recipients configured")
        return False
    
    try:
        msg = MIMEMultipart('mixed')  # Changed to 'mixed' for attachments
        msg['Subject'] = subject
        msg['From'] = f"SIEM Dashboard <{SMTP_EMAIL}>"
        msg['To'] = ', '.join(recipients)
        
        # Attach HTML content
        html_part = MIMEText(html_content, 'html', 'utf-8')
        msg.attach(html_part)
        
        # Attach Excel file if provided
        if excel_data and excel_filename:
            excel_part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_part.set_payload(excel_data)
            encoders.encode_base64(excel_part)
            excel_part.add_header('Content-Disposition', 'attachment', filename=excel_filename)
            msg.attach(excel_part)
            logger.info(f"Excel attachment added: {excel_filename}")
        
        # Send via SMTP
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, recipients, msg.as_string())
        
        logger.info(f"Email report sent successfully to {recipients}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False


def run_daily_off_hours_report():
    """
    Main function called by scheduler at 6:00 AM daily.
    Collects logs from 18:00 yesterday to 06:00 today and sends report.
    """
    logger.info("Starting daily off-hours report generation...")
    
    try:
        # Calculate time range: 18:00 yesterday -> 06:00 today
        now = datetime.now()
        to_time = now.replace(hour=6, minute=0, second=0, microsecond=0)
        from_time = (to_time - timedelta(days=1)).replace(hour=18, minute=0, second=0, microsecond=0)
        
        logger.info(f"Fetching logs from {from_time} to {to_time}")
        
        # Get logs
        logs = get_off_hours_logs(from_time, to_time)
        logger.info(f"Found {len(logs)} logs in off-hours period")
        
        # Generate report
        html_content = generate_html_report(logs, from_time, to_time)
        
        # Generate Excel attachment if logs > 20
        excel_data = None
        excel_filename = None
        if len(logs) > 20:
            excel_data = generate_excel_report(logs, from_time, to_time)
            excel_filename = f"off_hours_logs_{from_time.strftime('%Y%m%d')}.xlsx"
            logger.info(f"Generated Excel report: {excel_filename}")
        
        # Prepare subject
        stats = categorize_logs(logs)
        subject = f"🌙 Báo cáo ngoài giờ {from_time.strftime('%d/%m')} | {len(logs)} logs | {len(stats['violations'])} vi phạm"
        
        # Send email
        recipients = [r.strip() for r in REPORT_RECIPIENTS if r.strip()]
        if not recipients:
            logger.warning("No recipients configured, skipping email send")
            return
        
        success = send_email_report(subject, html_content, recipients, excel_data, excel_filename)
        
        if success:
            logger.info("Daily off-hours report sent successfully")
        else:
            logger.error("Failed to send daily off-hours report")
            
    except Exception as e:
        logger.error(f"Error in run_daily_off_hours_report: {e}")


def send_test_report(recipient_email: Optional[str] = None) -> Dict[str, Any]:
    """
    Send a test report for the last off-hours period.
    Used for testing email configuration.
    """
    try:
        # Use last night's off-hours period for test
        now = datetime.now()
        if now.hour < 6:
            # If before 6 AM, use yesterday 18:00 to now
            from_time = (now - timedelta(days=1)).replace(hour=18, minute=0, second=0, microsecond=0)
            to_time = now
        elif now.hour >= 18:
            # If after 6 PM, use today 18:00 to now
            from_time = now.replace(hour=18, minute=0, second=0, microsecond=0)
            to_time = now
        else:
            # During work hours, use last night
            from_time = (now - timedelta(days=1)).replace(hour=18, minute=0, second=0, microsecond=0)
            to_time = now.replace(hour=6, minute=0, second=0, microsecond=0)
        
        logs = get_off_hours_logs(from_time, to_time)
        html_content = generate_html_report(logs, from_time, to_time)
        
        # Generate Excel attachment
        excel_data = generate_excel_report(logs, from_time, to_time)
        excel_filename = f"test_off_hours_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        stats = categorize_logs(logs)
        subject = f"[TEST] 🌙 Báo cáo ngoài giờ | {len(logs)} logs"
        
        recipients = [recipient_email] if recipient_email else [r.strip() for r in REPORT_RECIPIENTS if r.strip()]
        
        success = send_email_report(subject, html_content, recipients, excel_data, excel_filename)
        
        return {
            'success': success,
            'logs_count': len(logs),
            'violations': len(stats['violations']),
            'warnings': len(stats['warnings']),
            'recipients': recipients,
            'time_range': f"{from_time} -> {to_time}",
            'excel_attached': True if excel_data else False
        }
        
    except Exception as e:
        logger.error(f"Error in send_test_report: {e}")
        return {'success': False, 'error': str(e)}
